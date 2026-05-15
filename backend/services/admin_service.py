
from datetime import datetime, timedelta, timezone
from typing import Optional
from fastapi import HTTPException
from core.database import db
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from core.constants import (
    MAX_LIMIT,
    ORDER_STATUS_CONFIRMED,
    ORDER_STATUS_DELIVERED,
    ORDER_STATUS_PACKED,
    ORDER_STATUS_SHIPPED,
    PAYMENT_STATUS_PAID,
)


LEGACY_REVENUE_STATUSES = [
    ORDER_STATUS_CONFIRMED,
    ORDER_STATUS_PACKED,
    ORDER_STATUS_SHIPPED,
    ORDER_STATUS_DELIVERED,
]


# Helper functions for dashboard-style date filtering (Excel export & dashboard)
def start_of_month(dt: datetime) -> datetime:
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def next_month_start(dt: datetime) -> datetime:
    if dt.month == 12:
        return dt.replace(year=dt.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    return dt.replace(month=dt.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)


def build_date_range(
    period: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None,
) -> tuple[Optional[datetime], Optional[datetime]]:
    now = datetime.now(timezone.utc)
    period = (period or "weekly").lower().strip()

    if period == "weekly":
        end_dt = now
        start_dt = now - timedelta(days=7)
        return start_dt, end_dt

    if period == "monthly":
        if month:
            try:
                month_dt = datetime.strptime(month, "%Y-%m").replace(tzinfo=timezone.utc)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
            start_dt = start_of_month(month_dt)
        else:
            start_dt = start_of_month(now)
        end_dt = next_month_start(start_dt)
        return start_dt, end_dt

    if period == "quarterly":
        quarter_start_month = ((now.month - 1) // 3) * 3 + 1
        start_dt = now.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        if quarter_start_month == 10:
            end_dt = start_dt.replace(year=start_dt.year + 1, month=1)
        else:
            end_dt = start_dt.replace(month=quarter_start_month + 3)
        return start_dt, end_dt

    if period == "yearly":
        start_dt = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = start_dt.replace(year=start_dt.year + 1)
        return start_dt, end_dt

    if period == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date are required for custom period")

        try:
            start_dt = datetime.fromisoformat(start_date)
            end_dt = datetime.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

        if start_dt.tzinfo is None:
            start_dt = start_dt.replace(tzinfo=timezone.utc)
        if end_dt.tzinfo is None:
            end_dt = end_dt.replace(tzinfo=timezone.utc)

        start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = end_dt.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
        return start_dt, end_dt

    if period == "all":
        return None, None

    raise HTTPException(status_code=400, detail="Invalid period")


def build_order_period_filter(
    period: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None,
) -> dict:
    start_dt, end_dt = build_date_range(period, start_date, end_date, month)
    if not start_dt or not end_dt:
        return {}

    return {
        "created_at": {
            "$gte": start_dt.isoformat(),
            "$lt": end_dt.isoformat(),
        }
    }


def build_revenue_order_filter() -> dict:
    return {
        "$or": [
            {"payment_status": PAYMENT_STATUS_PAID},
            {
                "payment_status": None,
                "status": {"$in": LEGACY_REVENUE_STATUSES},
            },
        ]
    }


def combine_filters(*filters: dict) -> dict:
    active_filters = [filter_item for filter_item in filters if filter_item]
    if not active_filters:
        return {}
    if len(active_filters) == 1:
        return active_filters[0]
    return {"$and": active_filters}


def is_revenue_eligible_order(order: dict) -> bool:
    if order.get("payment_status") == PAYMENT_STATUS_PAID:
        return True
    if order.get("payment_status") in (None, "") and order.get("status") in LEGACY_REVENUE_STATUSES:
        return True
    return False


def normalize_admin_payment_fields(order: dict) -> dict:
    if not order.get("payment_provider"):
        order["payment_provider"] = order.get("payment_method") or "manual_legacy"
    if not order.get("payment_status") and is_revenue_eligible_order(order):
        order["payment_status"] = PAYMENT_STATUS_PAID
    return order


async def get_dashboard_stats_service(
    period: str = "weekly",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None,
):
    period = (period or "weekly").lower().strip()
    match_filter = build_order_period_filter(period, start_date, end_date, month)
    revenue_filter = build_revenue_order_filter()
    period_revenue_filter = combine_filters(match_filter, revenue_filter)

    overall_revenue_pipeline = [
        {"$match": revenue_filter},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ]
    overall_revenue_result = await db.orders.aggregate(overall_revenue_pipeline).to_list(1)
    total_revenue = overall_revenue_result[0]["total"] if overall_revenue_result else 0

    period_revenue_pipeline = [
        {"$match": period_revenue_filter},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}},
    ]
    period_revenue_result = await db.orders.aggregate(period_revenue_pipeline).to_list(1)
    period_revenue = period_revenue_result[0]["total"] if period_revenue_result else 0

    total_paid_orders = await db.orders.count_documents(revenue_filter)
    period_paid_orders = await db.orders.count_documents(period_revenue_filter)
    total_operational_orders = await db.orders.count_documents({})
    period_operational_orders = await db.orders.count_documents(match_filter or {})
    total_products = await db.products.count_documents({})
    total_customers = await db.users.count_documents({"role": "user"})

    status_pipeline = []
    if match_filter:
        status_pipeline.append({"$match": match_filter})
    status_pipeline.append({"$group": {"_id": "$status", "count": {"$sum": 1}}})
    status_result = await db.orders.aggregate(status_pipeline).to_list(20)
    orders_by_status = {item["_id"]: item["count"] for item in status_result}

    recent_orders = await db.orders.find(match_filter or {}, {"_id": 0}).sort("created_at", -1).to_list(5)

    user_ids = list({order.get("user_id") for order in recent_orders if order.get("user_id")})
    users = await db.users.find(
        {"id": {"$in": user_ids}},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(len(user_ids) or 1)
    user_map = {user["id"]: user for user in users}

    for order in recent_orders:
        user = user_map.get(order.get("user_id"))
        if user:
            order["user_name"] = user.get("name", "")
            order["user_email"] = user.get("email", "")
        normalize_admin_payment_fields(order)

    if period == "yearly":
        group_id = {"$substr": ["$created_at", 0, 7]}
    else:
        group_id = {"$substr": ["$created_at", 0, 10]}

    period_pipeline = [{"$match": period_revenue_filter}]
    period_pipeline.extend([
        {
            "$group": {
                "_id": group_id,
                "orders": {"$sum": 1},
                "revenue": {"$sum": "$total_price"}
            }
        },
        {"$sort": {"_id": 1}}
    ])
    period_stats = await db.orders.aggregate(period_pipeline).to_list(366)

    start_dt, end_dt = build_date_range(period, start_date, end_date, month)

    return {
        "period": period,
        "month": month,
        "range_start": start_dt.isoformat() if start_dt else None,
        "range_end": end_dt.isoformat() if end_dt else None,
        "total_revenue": total_revenue,
        "total_orders": total_paid_orders,
        "total_paid_orders": total_paid_orders,
        "total_operational_orders": total_operational_orders,
        "total_products": total_products,
        "total_customers": total_customers,
        "period_revenue": period_revenue,
        "period_orders": period_paid_orders,
        "period_paid_orders": period_paid_orders,
        "period_operational_orders": period_operational_orders,
        "orders_by_status": orders_by_status,
        "recent_orders": recent_orders,
        "period_stats": period_stats,
        "weekly_stats": period_stats,
    }


async def export_orders_excel_service(
    period: str = "all",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None,
):
    match_filter = build_order_period_filter(period, start_date, end_date, month)

    orders = await db.orders.find(match_filter, {"_id": 0}).sort("created_at", -1).to_list(5000)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders Export"

    headers = [
        "Order ID",
        "Order Date",
        "Order Time",
        "Customer Name",
        "Customer Phone Number",
        "Customer Email ID",
        "Delivery Address",
        "City",
        "Pincode",
        "Product Name",
        "Product Category",
        "Quantity",
        "Price Per Unit",
        "Total Product Amount",
        "Delivery Charges",
        "Discount Applied",
        "Coupon Code",
        "Final Order Value",
        "Payment Method",
        "Payment Provider",
        "Payment Status",
        "Revenue Eligible",
        "Revenue Amount",
        "Order Status",
        "Delivery Date",
        "Delivery Time",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="1F1A17")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row_num = 2
    for order in orders:
        normalize_admin_payment_fields(order)
        created_at = order.get("created_at", "")
        order_date = ""
        order_time = ""
        if created_at:
            try:
                created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                order_date = created_dt.strftime("%Y-%m-%d")
                order_time = created_dt.strftime("%H:%M:%S")
            except Exception:
                order_date = created_at[:10]
                order_time = created_at[11:19] if len(created_at) >= 19 else ""

        items = order.get("items", []) or [{}]
        order_discount = order.get("discount_amount", order.get("discount", 0))
        coupon_code = order.get("coupon_code", "")
        delivery_charges = order.get("delivery_charges", 0)
        payment_status = order.get("payment_status", "")
        payment_provider = order.get("payment_provider", "")
        revenue_eligible = is_revenue_eligible_order(order)
        revenue_amount = order.get("total_price", 0) if revenue_eligible else 0
        delivery_date = order.get("delivery_date", "")
        delivery_time = order.get("delivery_time", "")

        for item in items:
            quantity = item.get("quantity", 0) or 0
            unit_price = item.get("price", 0) or 0
            total_product_amount = quantity * unit_price

            product_category = item.get("product_category") or item.get("category") or ""
            if not product_category and item.get("product_id"):
                product_doc = await db.products.find_one({"id": item.get("product_id")}, {"_id": 0, "category_id": 1})
                if product_doc and product_doc.get("category_id"):
                    category_doc = await db.categories.find_one({"id": product_doc["category_id"]}, {"_id": 0, "name": 1})
                    if category_doc:
                        product_category = category_doc.get("name", "")

            row = [
                order.get("id", ""),
                order_date,
                order_time,
                order.get("billing_name", ""),
                order.get("billing_phone", ""),
                order.get("billing_email", ""),
                order.get("billing_address", ""),
                order.get("billing_city", ""),
                order.get("billing_postal_code", ""),
                item.get("product_name", item.get("name", "")),
                product_category,
                quantity,
                unit_price,
                total_product_amount,
                delivery_charges,
                order_discount,
                coupon_code,
                order.get("total_price", 0),
                order.get("payment_method", ""),
                payment_provider,
                payment_status,
                "Yes" if revenue_eligible else "No",
                revenue_amount,
                order.get("status", ""),
                delivery_date,
                delivery_time,
            ]

            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_num, column=col_index, value=value)

            row_num += 1

    column_widths = {
        1: 16, 2: 14, 3: 12, 4: 24, 5: 18, 6: 28, 7: 34, 8: 18, 9: 12,
        10: 28, 11: 22, 12: 10, 13: 14, 14: 18, 15: 16, 16: 16, 17: 16,
        18: 18, 19: 16, 20: 18, 21: 16, 22: 16, 23: 16, 24: 16, 25: 14, 26: 14,
    }
    for col_index, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mariso_orders_{period}_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'}
    )

async def get_customers_service():
    users = await db.users.find({"role": "user"}, {"_id": 0, "password": 0}).to_list(MAX_LIMIT)
    
    for user in users:
        orders = await db.orders.find({"user_id": user['id']}, {"_id": 0}).to_list(MAX_LIMIT)
        user['total_orders'] = len(orders)
        user['total_spending'] = sum(order['total_price'] for order in orders)
    
    return users
